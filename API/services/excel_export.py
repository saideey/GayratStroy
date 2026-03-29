"""
Excel export uchun - Chiqimlar va Ta'minotchilar.
openpyxl ishlatiladi.
"""

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func, and_

from database.models.expense import (
    Expense, ExpenseCategory, ExpenseEditLog, ExpenseCurrencyType
)
from database.models.supplier import (
    Supplier, SupplierTransaction, SupplierTransactionType
)


# ── Stil konstantlari ─────────────────────────────────────────────────────────

def _make_header_fill(color="4472C4"):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _make_border():
    thin = Side(style='thin', color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=13)
BOLD_FONT   = Font(bold=True, size=11)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
RIGHT       = Alignment(horizontal="right",  vertical="center")

FILL_BLUE   = _make_header_fill("4472C4")
FILL_GREEN  = _make_header_fill("217346")
FILL_RED    = _make_header_fill("C0392B")
FILL_ORANGE = _make_header_fill("E67E22")
FILL_ALT    = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
FILL_DEL    = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
BORDER      = _make_border()


def _set_header(ws, row, cols, fill=FILL_BLUE):
    """Jadval sarlavhasini yozish."""
    for col_idx, (header, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _fmt_money(val) -> str:
    if val is None:
        return "0"
    return f"{int(float(val)):,}".replace(",", " ")


def _fmt_date(d) -> str:
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d.%m.%Y")
    return str(d)


def _fmt_datetime(d) -> str:
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d.%m.%Y %H:%M")
    return _fmt_date(d)


# ══════════════════════════════════════════════════════════════════════════════
# CHIQIMLAR EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generate_expenses_excel(
    db: Session,
    category_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    currency: Optional[str] = None,
    include_deleted: bool = False,
) -> bytes:
    """
    Chiqimlar Excel hisoboti.
    Sheet 1: Umumiy ro'yxat
    Sheet 2: Kategoriyalar bo'yicha xulosa
    Sheet 3: O'zgartirish tarixi (audit log)
    """
    wb = Workbook()

    # ── Sheet 1: Ro'yxat ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Chiqimlar"

    # Sarlavha
    ws1.merge_cells("A1:J1")
    title_cell = ws1["A1"]
    title_cell.value = "CHIQIMLAR HISOBOTI"
    title_cell.font = Font(bold=True, size=14, color="4472C4")
    title_cell.alignment = CENTER

    # Filter info
    ws1.merge_cells("A2:J2")
    filter_parts = []
    if start_date: filter_parts.append(f"Dan: {_fmt_date(start_date)}")
    if end_date:   filter_parts.append(f"Gacha: {_fmt_date(end_date)}")
    if currency:   filter_parts.append(f"Valyuta: {currency.upper()}")
    filter_str = " | ".join(filter_parts) if filter_parts else "Barcha"
    ws1["A2"].value = f"Filter: {filter_str}   |   Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="666666", size=10)
    ws1["A2"].alignment = LEFT

    # Jadval
    cols = [
        ("№",           5),
        ("Sana",        12),
        ("Sarlavha",    30),
        ("Kategoriya",  18),
        ("Summa",       15),
        ("Valyuta",     10),
        ("Kurs",        12),
        ("UZS ekviv.",  15),
        ("Kim yozdi",   18),
        ("Holat",       12),
    ]
    _set_header(ws1, 4, cols)

    # Qatorlar
    query = db.query(Expense)
    if not include_deleted:
        query = query.filter(Expense.is_deleted == False)
    if category_id:
        query = query.filter(Expense.category_id == category_id)
    if start_date:
        query = query.filter(Expense.expense_date >= start_date)
    if end_date:
        query = query.filter(Expense.expense_date <= end_date)
    if currency:
        query = query.filter(Expense.currency == ExpenseCurrencyType(currency))

    expenses = query.order_by(Expense.expense_date.desc(), Expense.id.desc()).all()

    total_uzs = Decimal("0")
    total_usd = Decimal("0")

    for i, e in enumerate(expenses, 1):
        row = 4 + i
        fill = FILL_DEL if e.is_deleted else (FILL_ALT if i % 2 == 0 else None)

        data = [
            i,
            _fmt_date(e.expense_date),
            e.title,
            e.category.name if e.category else "",
            float(e.amount),
            e.currency.value.upper(),
            float(e.usd_rate) if e.usd_rate else "",
            float(e.amount_uzs) if e.amount_uzs else float(e.amount),
            f"{e.created_by.first_name} {e.created_by.last_name}" if e.created_by else "",
            "O'chirilgan" if e.is_deleted else "Aktiv",
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = RIGHT if col_idx in (1, 5, 7, 8) else LEFT
            if fill:
                cell.fill = fill
            if e.is_deleted:
                cell.font = Font(color="999999", italic=True)

        if e.currency == ExpenseCurrencyType.UZS:
            total_uzs += e.amount
        else:
            total_usd += e.amount
            if e.amount_uzs:
                total_uzs += e.amount_uzs

    # Jami qator
    total_row = 4 + len(expenses) + 1
    ws1.merge_cells(f"A{total_row}:D{total_row}")
    ws1[f"A{total_row}"].value = "JAMI:"
    ws1[f"A{total_row}"].font = BOLD_FONT
    ws1[f"A{total_row}"].fill = _make_header_fill("17A589")
    ws1[f"A{total_row}"].font = Font(bold=True, color="FFFFFF")
    ws1[f"A{total_row}"].alignment = RIGHT

    ws1[f"H{total_row}"].value = float(total_uzs)
    ws1[f"H{total_row}"].font = Font(bold=True, color="C0392B")
    ws1[f"H{total_row}"].fill = _make_header_fill("FDECEA")
    ws1[f"H{total_row}"].number_format = '#,##0'

    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 18
    ws1.row_dimensions[4].height = 25

    # ── Sheet 2: Kategoriyalar ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Kategoriyalar bo'yicha")

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "KATEGORIYALAR BO'YICHA XULOSA"
    ws2["A1"].font = Font(bold=True, size=13, color="4472C4")
    ws2["A1"].alignment = CENTER

    cat_cols = [
        ("Kategoriya",   25),
        ("Chiqimlar soni", 18),
        ("UZS jami",     18),
        ("USD jami",     15),
        ("UZS ekvival.", 18),
    ]
    _set_header(ws2, 3, cat_cols, fill=FILL_GREEN)

    # Kategoriya bo'yicha hisob
    from sqlalchemy import case as sa_case
    cat_data = db.query(
        ExpenseCategory.name,
        func.count(Expense.id).label("cnt"),
        func.sum(
            sa_case((Expense.currency == ExpenseCurrencyType.UZS, Expense.amount), else_=0)
        ).label("uzs_total"),
        func.sum(
            sa_case((Expense.currency == ExpenseCurrencyType.USD, Expense.amount), else_=0)
        ).label("usd_total"),
        func.coalesce(func.sum(Expense.amount_uzs), 0).label("uzs_equiv"),
    ).join(
        Expense, and_(
            Expense.category_id == ExpenseCategory.id,
            Expense.is_deleted == False
        ), isouter=True
    ).filter(ExpenseCategory.is_active == True).group_by(
        ExpenseCategory.id, ExpenseCategory.name
    ).order_by(func.sum(Expense.amount_uzs).desc().nullslast()).all()

    for i, row_data in enumerate(cat_data, 1):
        row = 3 + i
        fill = FILL_ALT if i % 2 == 0 else None
        data = [
            row_data.name,
            row_data.cnt or 0,
            float(row_data.uzs_total or 0),
            float(row_data.usd_total or 0),
            float(row_data.uzs_equiv or 0),
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = RIGHT if col_idx > 1 else LEFT
            if fill:
                cell.fill = fill
            if col_idx in (3, 4, 5):
                cell.number_format = '#,##0'

    # ── Sheet 3: Audit log ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("O'zgartirish tarixi")

    ws3.merge_cells("A1:H1")
    ws3["A1"].value = "CHIQIMLAR O'ZGARTIRISH TARIXI"
    ws3["A1"].font = Font(bold=True, size=13, color="4472C4")
    ws3["A1"].alignment = CENTER

    audit_cols = [
        ("Vaqt",         17),
        ("Amal",         12),
        ("Chiqim ID",    10),
        ("Kim",          18),
        ("Izoh",         35),
        ("Eski summa",   14),
        ("Yangi summa",  14),
        ("Eski kateg.",  16),
    ]
    _set_header(ws3, 3, audit_cols, fill=FILL_ORANGE)

    ACTION_LABELS = {
        "created": "Yaratildi",
        "updated": "Yangilandi",
        "deleted": "O'chirildi",
        "restored": "Tiklandi",
    }
    ACTION_FILLS = {
        "created":  _make_header_fill("E8F8F5"),
        "updated":  _make_header_fill("EBF5FB"),
        "deleted":  FILL_DEL,
        "restored": _make_header_fill("FEF9E7"),
    }

    logs = db.query(ExpenseEditLog).order_by(
        ExpenseEditLog.created_at.desc()
    ).limit(500).all()

    for i, log in enumerate(logs, 1):
        row = 3 + i
        action = log.action.value if log.action else ""
        fill = ACTION_FILLS.get(action, None)
        data = [
            _fmt_datetime(log.created_at),
            ACTION_LABELS.get(action, action),
            log.expense_id,
            f"{log.changed_by.first_name} {log.changed_by.last_name}" if log.changed_by else "",
            log.comment,
            float(log.old_amount) if log.old_amount else "",
            float(log.new_amount) if log.new_amount else "",
            log.old_category_id or "",
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws3.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = LEFT
            if fill:
                cell.fill = fill

    # Saqlash
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# TA'MINOTCHILAR EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generate_suppliers_excel(
    db: Session,
    supplier_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    transaction_type: Optional[str] = None,
) -> bytes:
    """
    Ta'minotchilar Excel hisoboti.
    supplier_id berilsa — bitta ta'minotchi, aks holda hammasi.
    Sheet 1: Ta'minotchilar ro'yxati (balans)
    Sheet 2: Tranzaksiyalar tarixi
    """
    wb = Workbook()

    # ── Sheet 1: Ta'minotchilar ro'yxati ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ta'minotchilar"

    ws1.merge_cells("A1:I1")
    ws1["A1"].value = "TA'MINOTCHILAR HISOBOTI"
    ws1["A1"].font = Font(bold=True, size=14, color="4472C4")
    ws1["A1"].alignment = CENTER

    ws1.merge_cells("A2:I2")
    ws1["A2"].value = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="666666", size=10)
    ws1["A2"].alignment = LEFT

    sup_cols = [
        ("№",            5),
        ("Nomi",         25),
        ("Kompaniya",    22),
        ("Telefon",      15),
        ("Shahar",       12),
        ("INN",          13),
        ("Joriy qarz",   15),
        ("Jami qarz*",   15),
        ("Jami to'lov*", 15),
    ]
    _set_header(ws1, 4, sup_cols)

    sup_query = db.query(Supplier).filter(Supplier.is_deleted == False)
    if supplier_id:
        sup_query = sup_query.filter(Supplier.id == supplier_id)
    suppliers = sup_query.order_by(Supplier.current_debt.desc(), Supplier.name).all()

    grand_debt = Decimal("0")
    grand_total_debt = Decimal("0")
    grand_total_paid = Decimal("0")

    for i, s in enumerate(suppliers, 1):
        row = 4 + i
        fill = FILL_ALT if i % 2 == 0 else None

        # Tranzaksiyalardan hisoblash
        tx_query = db.query(SupplierTransaction).filter(
            SupplierTransaction.supplier_id == s.id,
            SupplierTransaction.is_deleted == False
        )
        debt_sum = db.query(func.coalesce(func.sum(SupplierTransaction.amount_uzs), 0)).filter(
            SupplierTransaction.supplier_id == s.id,
            SupplierTransaction.is_deleted == False,
            SupplierTransaction.transaction_type == SupplierTransactionType.DEBT
        ).scalar() or Decimal("0")

        paid_sum = db.query(func.coalesce(func.sum(SupplierTransaction.amount_uzs), 0)).filter(
            SupplierTransaction.supplier_id == s.id,
            SupplierTransaction.is_deleted == False,
            SupplierTransaction.transaction_type.in_([
                SupplierTransactionType.PAYMENT, SupplierTransactionType.RETURN
            ])
        ).scalar() or Decimal("0")

        grand_debt += (s.current_debt or Decimal("0"))
        grand_total_debt += debt_sum
        grand_total_paid += paid_sum

        data = [
            i,
            s.name,
            s.company_name or "",
            s.phone or "",
            s.city or "",
            s.inn or "",
            float(s.current_debt or 0),
            float(debt_sum),
            float(paid_sum),
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = RIGHT if col_idx in (1, 7, 8, 9) else LEFT
            if fill:
                cell.fill = fill
            if col_idx == 7 and float(s.current_debt or 0) > 0:
                cell.font = Font(bold=True, color="C0392B")
            if col_idx in (7, 8, 9):
                cell.number_format = '#,##0'

    # Jami qator
    total_row = 4 + len(suppliers) + 1
    ws1.merge_cells(f"A{total_row}:F{total_row}")
    ws1[f"A{total_row}"].value = "JAMI:"
    ws1[f"A{total_row}"].font = Font(bold=True, color="FFFFFF")
    ws1[f"A{total_row}"].fill = _make_header_fill("17A589")
    ws1[f"A{total_row}"].alignment = RIGHT

    for col, val in [(7, float(grand_debt)), (8, float(grand_total_debt)), (9, float(grand_total_paid))]:
        cell = ws1.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, color="C0392B" if col == 7 else "1A5276")
        cell.fill = _make_header_fill("FDECEA" if col == 7 else "EBF5FB")
        cell.number_format = '#,##0'
        cell.border = BORDER

    # Izoh
    note_row = total_row + 1
    ws1[f"A{note_row}"].value = "* — Tranzaksiyalar tarixidan hisoblab chiqilgan"
    ws1[f"A{note_row}"].font = Font(italic=True, color="888888", size=9)

    # ── Sheet 2: Tranzaksiyalar ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Tranzaksiyalar")

    ws2.merge_cells("A1:J1")
    ws2["A1"].value = "TRANZAKSIYALAR TARIXI"
    ws2["A1"].font = Font(bold=True, size=13, color="4472C4")
    ws2["A1"].alignment = CENTER

    filter_info = []
    if supplier_id:
        sup = db.query(Supplier).filter(Supplier.id == supplier_id).first()
        if sup:
            filter_info.append(f"Ta'minotchi: {sup.name}")
    if start_date: filter_info.append(f"Dan: {_fmt_date(start_date)}")
    if end_date:   filter_info.append(f"Gacha: {_fmt_date(end_date)}")
    if transaction_type: filter_info.append(f"Tur: {transaction_type}")
    ws2.merge_cells("A2:J2")
    ws2["A2"].value = " | ".join(filter_info) if filter_info else "Barcha tranzaksiyalar"
    ws2["A2"].font = Font(italic=True, color="666666", size=10)
    ws2["A2"].alignment = LEFT

    tx_cols = [
        ("№",             5),
        ("Sana",         12),
        ("Ta'minotchi",  25),
        ("Tur",          14),
        ("Summa",        15),
        ("Valyuta",       9),
        ("UZS ekviv.",   15),
        ("Izoh",         35),
        ("Kim yozdi",    18),
        ("Holat",        12),
    ]
    _set_header(ws2, 4, tx_cols)

    TX_LABELS = {
        "debt":    "Qarz",
        "payment": "To'lov",
        "return":  "Qaytarish",
    }
    TX_FILLS = {
        "debt":    _make_header_fill("FDEDEC"),
        "payment": _make_header_fill("E9F7EF"),
        "return":  _make_header_fill("EBF5FB"),
    }

    tx_query = db.query(SupplierTransaction)
    if supplier_id:
        tx_query = tx_query.filter(SupplierTransaction.supplier_id == supplier_id)
    if start_date:
        tx_query = tx_query.filter(SupplierTransaction.transaction_date >= start_date)
    if end_date:
        tx_query = tx_query.filter(SupplierTransaction.transaction_date <= end_date)
    if transaction_type:
        tx_query = tx_query.filter(
            SupplierTransaction.transaction_type == SupplierTransactionType(transaction_type)
        )

    txs = tx_query.order_by(
        SupplierTransaction.transaction_date.desc(),
        SupplierTransaction.id.desc()
    ).all()

    for i, tx in enumerate(txs, 1):
        row = 4 + i
        tx_type = tx.transaction_type.value
        fill = FILL_DEL if tx.is_deleted else TX_FILLS.get(tx_type, None)
        data = [
            i,
            _fmt_date(tx.transaction_date),
            tx.supplier.name if tx.supplier else "",
            TX_LABELS.get(tx_type, tx_type),
            float(tx.amount),
            tx.currency.upper(),
            float(tx.amount_uzs) if tx.amount_uzs else float(tx.amount),
            tx.comment,
            f"{tx.created_by.first_name} {tx.created_by.last_name}" if tx.created_by else "",
            "O'chirilgan" if tx.is_deleted else "Aktiv",
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = RIGHT if col_idx in (1, 5, 7) else LEFT
            if fill:
                cell.fill = fill
            if tx.is_deleted:
                cell.font = Font(color="999999", italic=True)
            if col_idx in (5, 7):
                cell.number_format = '#,##0'

    # Saqlash
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
